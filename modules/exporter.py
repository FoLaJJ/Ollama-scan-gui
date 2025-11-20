# -*- coding: utf-8 -*-
"""
结果导出模块
支持导出为CSV, JSON, Excel格式
"""

import csv
import json
import os
from typing import List

import openpyxl
from openpyxl.styles import Font, PatternFill



class ResultExporter:
    """结果导出器"""
    
    @staticmethod
    def export(results: List[dict], file_path: str, format_type: str = "csv") -> bool:
        """
        导出结果
        
        Args:
            results: 结果列表，每个元素为字典
            file_path: 导出文件路径
            format_type: 导出格式 (csv, json, excel)
            
        Returns:
            bool: 是否成功
        """
        try:
            # 确保目录存在
            directory = os.path.dirname(file_path)
            if directory and not os.path.exists(directory):
                os.makedirs(directory)
            
            if format_type.lower() == "csv":
                return ResultExporter._export_csv(results, file_path)
            elif format_type.lower() == "json":
                return ResultExporter._export_json(results, file_path)
            elif format_type.lower() in ["excel", "xlsx"]:
                return ResultExporter._export_excel(results, file_path)
            else:
                raise ValueError(f"不支持的导出格式: {format_type}")
        
        except Exception as e:
            print(f"导出失败: {str(e)}")
            return False
    
    @staticmethod
    def _export_csv(results: List[dict], file_path: str) -> bool:
        """导出为CSV格式"""
        if not results:
            return False
        
        # 确保文件扩展名正确
        if not file_path.endswith('.csv'):
            file_path += '.csv'
        
        # 获取所有字段
        fieldnames = list(results[0].keys())
        
        with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(results)
        
        return True
    
    @staticmethod
    def _export_json(results: List[dict], file_path: str) -> bool:
        """导出为JSON格式"""
        if not results:
            return False
        
        # 确保文件扩展名正确
        if not file_path.endswith('.json'):
            file_path += '.json'
        
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(results, f, ensure_ascii=False, indent=2)
        
        return True
    
    @staticmethod
    def _export_excel(results: List[dict], file_path: str) -> bool:
        """导出为Excel格式"""
        if not results:
            return False
        
        # 确保文件扩展名正确
        if not file_path.endswith('.xlsx'):
            file_path += '.xlsx'
        
        # 创建工作簿
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "扫描结果"
        
        # 获取所有字段
        fieldnames = list(results[0].keys())
        
        # 写入表头
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_idx, field in enumerate(fieldnames, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=field)
            cell.fill = header_fill
            cell.font = header_font
        
        # 写入数据
        for row_idx, result in enumerate(results, start=2):
            for col_idx, field in enumerate(fieldnames, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=result.get(field, ""))
        
        # 自动调整列宽
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        workbook.save(file_path)
        workbook.close()
        
        return True
    
    @staticmethod
    def filter_results(results: List[dict], start: int = 0, 
                      count: int = None, vulnerable_only: bool = False) -> List[dict]:
        """
        过滤结果
        
        Args:
            results: 原始结果列表
            start: 起始索引（从0开始）
            count: 导出数量，None表示全部
            vulnerable_only: 是否只导出有漏洞的结果
            
        Returns:
            List[dict]: 过滤后的结果
        """
        # 过滤有漏洞的结果
        if vulnerable_only:
            results = [r for r in results if r.get("vulnerable", False)]
        
        # 切片
        if count is None:
            return results[start:]
        else:
            return results[start:start + count]
